import numpy as np
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings


class DataProcessor:
    def __init__(self):
        warnings.filterwarnings("ignore")
        self.setup_paths()

    def setup_paths(self):
        current_directory = os.getcwd()
        data_folder = os.path.join(current_directory, "data")
        self.input_path = os.path.join(data_folder, "dashboard_input.xlsx")
        self.output_path = os.path.join(data_folder, "dashboard_output.xlsx")
        self.inc_opps_path = os.path.join(data_folder, "query_output.xlsx")

    def load_data(self):
        print("\n" + "*" * 15 + "  Starting to Analyze  " + "*" * 15 + "\n")
        self.avn_df = pd.read_excel(self.input_path, "avn")
        self.aon_df = pd.read_excel(self.input_path, "aon")
        self.offline_df = pd.read_excel(self.input_path, "offline")
        self.inc_opps = pd.read_excel(self.inc_opps_path, "VM-Vendor Level")
        return self.aon_df, self.avn_df, self.offline_df, self.inc_opps


class DataCleaner:
    def clean_dataframes(self, aon_df, avn_df, offline_df):
        # Select and rename columns
        aon_selected_cols = ["Comp Code", "Prod Line", "Negotiator", "Total Neg Gains"]
        avn_selected_cols = ["GL", "Company Code", "USD Net Gains + LOS Gains (USD)"]
        
        aon_df = aon_df[aon_selected_cols]
        avn_df = avn_df[avn_selected_cols]

        # Rename columns
        aon_df.rename(columns={"Total Neg Gains": "Dashboard Gains"}, inplace=True)
        avn_df.rename(
            columns={
                "USD Net Gains + LOS Gains (USD)": "AVN Gains",
                "Company Code": "Comp Code",
                "GL": "Prod Line",
            },
            inplace=True,
        )
        offline_df.rename(
            columns={
                "Company Code": "Comp Code",
                "GL": "Prod Line",
                "Coop Amount": "Offline Gains",
            },
            inplace=True,
        )

        return aon_df, avn_df, offline_df


class DataAnalyzer:
    def create_avn_aon_view(self, aon_df, avn_df):
        avn_aon_df = aon_df.merge(avn_df, how="right", on=["Comp Code", "Prod Line"])
        avn_aon_df.fillna(0, inplace=True)

        avn_aon_df["AON Gains"] = np.where(
            avn_aon_df["Dashboard Gains"] > 0,
            avn_aon_df["Dashboard Gains"] - avn_aon_df["AVN Gains"],
            0,
        )
        return avn_aon_df

    def create_aon_gl_view(self, avn_aon_df):
        return avn_aon_df[["Prod Line", "AON Gains"]].groupby("Prod Line").sum().reset_index()

    def create_vm_vendor_view(self, avn_aon_df):
        vm_vendor = avn_aon_df[['Prod Line', 'Negotiator', 'Comp Code', 'AVN Gains']]
        vm_vendor.rename(columns={'AVN Gains':'Reported AVN Gains'}, inplace=True)
        
        return vm_vendor

    def create_inc_opps_views(self, inc_opps, aon_df_full ):
        ''' for GL level check in (vm level) '''
        # vm_vendor = self.create_vm_vendor_view(avn_aon_df)
        dashboard_inc_deal = aon_df_full[['Comp Code', 'Comp Name', 'Prod Line', 'Promotions']]
        dashboard_inc_deal.rename(columns={'Comp Code':'company_code', "Comp Name":"company_name",  'Prod Line':'prod_line', 'Promotions':'dashboard_promo'}, inplace=True)
        inc_opps = inc_opps[['prod_line',  'company_code', 'incremental_gains']]
        inc_opps = inc_opps[inc_opps['incremental_gains'] > 0]
        
        # mergee        
        vm_opps = dashboard_inc_deal .merge(
            inc_opps, 
            how='left', 
            on=['prod_line','company_code']
        )
        vm_opps['inc_deal_opps'] = vm_opps['incremental_gains'] - vm_opps['dashboard_promo']
        vm_opps = vm_opps[[
            'prod_line', 'company_code', 'company_name', 
            'incremental_gains', 'dashboard_promo', 'inc_deal_opps'
        ]]
        vm_opps.fillna(0, inplace=True)
        vm_opps = vm_opps[vm_opps['inc_deal_opps'] >= 0]
        vm_opps.sort_values(by='inc_deal_opps', ascending=False, inplace=True)
        return vm_opps

class ExcelWriter:
    def write_to_excel(self, outputs, output_path):
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

        for sheet_name, data in outputs.items():
            ws = wb.create_sheet(sheet_name)
            rows = dataframe_to_rows(data, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        print(f"\nOutput saved to {output_path}\n")


def main():
    # Initialize processor and load data
    processor = DataProcessor()
    aon_df, avn_df, offline_df, inc_opps = processor.load_data()
    aon_df_full = aon_df.copy()
    
    # Clean data
    cleaner = DataCleaner()
    aon_df, avn_df, offline_df = cleaner.clean_dataframes(aon_df, avn_df, offline_df)

    # Analyze data
    analyzer = DataAnalyzer()
    avn_aon_df = analyzer.create_avn_aon_view(aon_df, avn_df)
    aon_gl = analyzer.create_aon_gl_view(avn_aon_df)
    vm_vendor = analyzer.create_vm_vendor_view(avn_aon_df)
    vm_opps = analyzer.create_inc_opps_views(inc_opps, aon_df_full )  

    # Prepare outputs
    outputs = {
        "aon_vendor": avn_aon_df,
        "aon_gl": aon_gl,
        "vm_vendor": vm_vendor,
        "vm_opps": vm_opps,
    }

    # Write to Excel
    writer = ExcelWriter()
    writer.write_to_excel(outputs, processor.output_path)

    return outputs

if __name__ == "__main__":
    main()
