import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import datetime


class DataProcessor:
    def __init__(self):
        warnings.filterwarnings("ignore")
        self.setup_paths()

    def setup_paths(self):
        current_directory = os.getcwd()
        data_folder = os.path.join(current_directory, "data")
        self.query_input_path = os.path.join(data_folder, "query_input.csv")
        self.polo_path = os.path.join(data_folder, "polo.csv")
        self.output_path = os.path.join(data_folder, "query_output.xlsx")

    def load_data(self):
        print("\n" + "*" * 15 + "  Starting to Analyze  " + "*" * 15 + "\n")
        self.query_df = pd.read_csv(self.query_input_path)
        self.polo_df = pd.read_csv(self.polo_path)
        return self.query_df, self.polo_df

    def merge_data(self):
        self.query_df, self.polo_df = self.load_data()

        self.polo_df = self.polo_df[["userid", "name", "gl", "companyCode"]]
        polo_first_owner = self.polo_df.groupby(['gl', 'companyCode']).first().reset_index()
        polo_first_owner.rename(
            columns={
                "userid": "vm_alias",
                "name": "vm_name",
                "gl": "product_group_id",
                "companyCode": "company_code",
            },
            inplace=True,
        )
        self.df = self.query_df.merge(
            polo_first_owner,
            how="left",
            on=["product_group_id", "company_code"],
        )
        return self.df
    

class DataCleaner:
    def convert_dtypes(self, df):
        df["start_datetime"] = pd.to_datetime(
            df["start_datetime"], format="mixed", errors="coerce"
        )
        df["end_datetime"] = pd.to_datetime(
            df["end_datetime"], format="mixed", errors="coerce"
        )

        df["paws_promotion_id"] = np.where(
            (df["paws_promotion_id"] == np.nan)
            | (df["paws_promotion_id"].isnull())
            | (df["paws_promotion_id"] == "n/a")
            | (df["paws_promotion_id"] == "n/a"),
            0,
            df["paws_promotion_id"],
        )

        df["paws_promotion_id"] = df["paws_promotion_id"].astype(str)
        return df

    def standardize_columns(self, df):
        df.rename(columns={"quantity_sum": "shipped_units"}, inplace=True)
        df["start_year_mo"] = df["start_datetime"].dt.to_period("M")
        df.rename(columns={"start_year_mo": "reporting_period"}, inplace=True)

        float_cols = ["t4w_asp", "promotion_pricing_amount", "coop_amount"]
        int_cols = [
            "shipped_units",
            "region_id",
            "marketplace_key",
            "product_group_id",
            "agreement_id",
        ]
        str_cols = ["reporting_period", "vendor_code", "vendor_name"]

        for col in float_cols:
            df[col] = df[col].astype(float)
        for col in int_cols:
            df[col] = df[col].astype(int)
        for col in str_cols:
            df[col] = df[col].astype(str)

        df.rename(
            columns={
                "start_datetime": "promo_start_datetime",
                "end_datetime": "promo_end_datetime",
            },
            inplace=True,
        )
        return df

    def map_gl_codes(self, df):
        
        gl_mapping = {
            "121": "Health & Personal Care",
            "194": "Beauty",
            "199": "Pets",
            "325": "Grocery",
            "364": "Health & Personal Care",
            "510": "Lux Beauty",
            "75": "Baby",
        }

        df["product_group_id"] = df["product_group_id"].astype(str)

        for gl_code, category in gl_mapping.items():
            df["product_group_id"] = np.where(
                df["product_group_id"] == gl_code, category, df["product_group_id"]
            )
            
        df.rename(columns={"product_group_id": "prod_line"}, inplace=True)
        
        return df

    def convert_to_usd(self, df):
        # USD:CAD exchange rate
        USD_CAD_RATE = 1.43832
        
        # Columns containing monetary values
        monetary_columns = ['t4w_asp', 'promotion_pricing_amount', 'coop_amount']
        
        # Convert each monetary column from CAD to USD
        for col in monetary_columns:
            if col in df.columns:
                df[col] = df[col] / USD_CAD_RATE
                df[col] = round(df[col], 2)
        
        return df

class DataAggregator:
    def get_group_columns(self):
        return [
            "asin",
            "reporting_period",
            "promo_start_datetime",
            "promo_end_datetime",
            "region_id",
            "marketplace_key",
            "promotion_key",
            "paws_promotion_id",
            "vm_alias",
            "vm_name",
            "purpose",
            "agreement_id",
            "funding_type_name",
            "activity_type_name",
            "prod_line",
            "company_code",
            "vendor_code",
            "vendor_name",
            "t4w_asp",
            "promotion_pricing_amount",
            "funding_per_asin",
        ]

    def get_final_column_order(self):
        return [
            "asin",
            "reporting_period",
            "promo_start_datetime",
            "promo_end_datetime",
            "region_id",
            "marketplace_key",
            "promotion_key",
            "paws_promotion_id",
            "agreement_id",
            "vm_alias",
            "vm_name",
            "purpose",
            "funding_type_name",
            "activity_type_name",
            "prod_line",
            "company_code",
            "vendor_code",
            "vendor_name",
            "t4w_asp",
            "promotion_pricing_amount",
            "discount_per_unit",
            "funding_per_asin",
            "incremental_per_unit",
            "shipped_units",
            "incremental_gains",
        ]

    def aggregate_to_asin_level(self, df):
        df["funding_per_asin"] = df["coop_amount"] / df["shipped_units"]
        df["funding_per_asin"] = round(df["funding_per_asin"], 2)
        df.drop(columns=["coop_amount"], inplace=True)

        group_cols = self.get_group_columns()
        asin_level_df = (
            df.groupby(group_cols).agg({"shipped_units": "sum"}).reset_index()
        )

        return asin_level_df

    def create_vendor_agreement_view(self, df):
        vendor_cols = [
            "region_id",
            "marketplace_key",
            "reporting_period",
            "prod_line",
            "vendor_code",
            "vm_alias",
            "vm_name",
            "agreement_id",
            "funding_per_asin",
            "incremental_gains",
        ]

        vendor_agreement_df = (
            df[vendor_cols]
            .groupby(vendor_cols[:-1])
            .sum("incremental_gains")
            .reset_index()
        )

        vendor_agreement_df.sort_values(
            ["incremental_gains"], ascending=False, inplace=True
        )

        return vendor_agreement_df

    def create_vm_vendor_view(self, df):
        group_cols = [
            "region_id",
            "marketplace_key",
            "prod_line",
            "vm_alias",
            "vm_name",
            "company_code",
            # "vendor_name",
            "incremental_gains",
        ]

        res = (
            df[group_cols]
            .groupby(group_cols[:-1])
            .sum("incremental_gains")
            .reset_index()
        )

        res.sort_values(["incremental_gains"], ascending=False, inplace=True)

        return res

    def create_gl_view(self, df):
        group_cols = [
            "region_id",
            "marketplace_key",
            "prod_line",
            "incremental_gains",
        ]

        res = (
            df[group_cols]
            .groupby(group_cols[:-1])
            .sum("incremental_gains")
            .reset_index()
        )

        res.sort_values(["incremental_gains"], ascending=False, inplace=True)

        return res


class Calculator:
    def calculate_incremental_metrics(self, df):
        # Calculate discount per unit
        df["discount_per_unit"] = df["t4w_asp"] - df["promotion_pricing_amount"]
        df["discount_per_unit"] = np.where(
            df["t4w_asp"].isna(), np.nan, df["discount_per_unit"]
        )
        df["discount_per_unit"] = round(df["discount_per_unit"], 2)

        # filter out null rows
        df = df[(df["discount_per_unit"] > 0) & (df["funding_per_asin"] > 0)]

        # Calculate incremental per unit
        df["incremental_per_unit"] = df["funding_per_asin"] - df["discount_per_unit"]
        df["incremental_per_unit"] = np.where(
            (df["incremental_per_unit"] < 0) | (df["incremental_per_unit"].isna()),
            0,
            df["incremental_per_unit"],
        )
        df["incremental_per_unit"] = round(df["incremental_per_unit"], 2)

        # Calculate total incremental gains
        df["incremental_gains"] = df["incremental_per_unit"] * df["shipped_units"]
        df["incremental_gains"] = np.where(
            (df["incremental_gains"] < 0) | (df["incremental_gains"].isna()),
            0,
            df["incremental_gains"],
        )
        df["incremental_gains"] = round(df["incremental_gains"], 2)

        return df

    def sort_and_order_columns(self, df):
        # Sort values
        df = df.sort_values(
            by=["marketplace_key", "region_id", "incremental_gains"],
            ascending=[True, True, False],
        )

        # Reorder columns
        aggregator = DataAggregator()
        col_order = aggregator.get_final_column_order()
        df = df[col_order]

        return df


class ExcelWriter:
    def write_to_excel(self, df, vendor_agreement_df, vm_vendor_df, gl_df, output_path):
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

        outputs = {
            "ASIN Level": df,
            "Vendor-Agreement Level": vendor_agreement_df,
            "VM-Vendor Level": vm_vendor_df,
            "GL Level": gl_df,
        }

        for sheet_name, data in outputs.items():
            ws = wb.create_sheet(sheet_name)
            rows = dataframe_to_rows(data, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        print("\n" + "*" * 8 + f"  Output saved to {output_path}.  " + "*" * 8 + "\n")

def main():
    # Initialize processor and load data
    processor = DataProcessor()
    df = processor.merge_data()
    df.dropna(inplace=True, how="all")

    # Clean data
    cleaner = DataCleaner()
    # Convert CAD to USD first
    df = cleaner.convert_to_usd(df)
    # Continue with other cleaning steps
    df = cleaner.convert_dtypes(df)
    df = cleaner.standardize_columns(df)
    df = cleaner.map_gl_codes(df)


    # Aggregate to ASIN level first
    aggregator = DataAggregator()
    asin_level_df = aggregator.aggregate_to_asin_level(df)

    # Calculate incremental metrics
    calculator = Calculator()
    asin_level_df = calculator.calculate_incremental_metrics(asin_level_df)
    asin_level_df = calculator.sort_and_order_columns(asin_level_df)

    # Create vendor agreement view
    vendor_agreement_df = aggregator.create_vendor_agreement_view(asin_level_df)

    # Create VM vendor view
    vm_vendor_df = aggregator.create_vm_vendor_view(asin_level_df)

    # create GL level view
    gl_df = aggregator.create_gl_view(asin_level_df)

    # Write to Excel
    writer = ExcelWriter()
    writer.write_to_excel(
        asin_level_df, vendor_agreement_df, vm_vendor_df, gl_df, processor.output_path
    )
    
    dct = {
        "asin_level": asin_level_df,
        "vendor_agreement": vendor_agreement_df,
        "vm_vendor": vm_vendor_df,
        "gl_level": gl_df,
    }
    return dct

if __name__ == "__main__":
    main()
