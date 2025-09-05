import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings


class DataProcessor:
    def __init__(self):
        warnings.filterwarnings("ignore")
        self.setup_paths()

    def setup_paths(self):
        current_directory = os.getcwd()
        # base_folder = os.path.dirname(current_directory)
        data_folder = os.path.join(current_directory, "data")
        self.input_path = os.path.join(data_folder, "input.xlsx")
        self.output_path = os.path.join(data_folder, "output.xlsx")

    def load_data(self):
        print("\n" + "*" * 15 + "  Starting to Analyze  " + "*" * 15 + "\n")
        self.df = pd.read_excel(self.input_path)
        return self.df


class DataCleaner:
    @staticmethod
    def convert_dtypes(df):
        # Convert datetime columns
        df["start_datetime"] = pd.to_datetime(
            df["start_datetime"], format="%d%b%Y:%H:%M:%S.%f"
        )
        df["end_datetime"] = pd.to_datetime(
            df["end_datetime"], format="%d%b%Y:%H:%M:%S.%f"
        )

        # Handle NA and inf values
        df["paws_promotion_id"] = np.where(
            (df["paws_promotion_id"] == np.nan)
            | (df["paws_promotion_id"].isnull())
            | (df["paws_promotion_id"] == np.inf)
            | (df["paws_promotion_id"] == -np.inf),
            "",
            df["paws_promotion_id"].astype(int),
        )
        df["paws_promotion_id"] = df["paws_promotion_id"].astype(int)
        return df

    @staticmethod
    def standardize_columns(df):
        # Rename columns
        df.rename(columns={"quantity_sum": "shipped_units"}, inplace=True)
        df["start_year_mo"] = df["start_datetime"].dt.to_period("M")
        df.rename(columns={"start_year_mo": "period"}, inplace=True)

        # Convert datatypes
        float_cols = ["t4w_asp", "promotion_pricing_amount", "coop_amount"]
        int_cols = [
            "shipped_units",
            "region_id",
            "marketplace_key",
            "paws_promotion_id",
            "product_group_id",
        ]
        str_cols = ["period", "vendor_code", "company_code", "company_name"]

        for col in float_cols:
            df[col] = df[col].astype(float)
        for col in int_cols:
            df[col] = df[col].astype(int)
        for col in str_cols:
            df[col] = df[col].astype(str)

        return df


class DataAggregator:
    @staticmethod
    def get_group_columns():
        return [
            "asin",
            "period",
            "region_id",
            "marketplace_key",
            "promotion_key",
            "paws_promotion_id",
            "owned_by_user_id",
            "purpose",
            "funding_type_name",
            "activity_type_name",
            "product_group_id",
            "vendor_code",
            "company_code",
            "company_name",
            "t4w_asp",
            "promotion_pricing_amount",
            "funding_per_asin",
        ]

    @staticmethod
    def get_final_column_order():
        return [
            "asin",
            "period",
            "region_id",
            "marketplace_key",
            "promotion_key",
            "paws_promotion_id",
            "owned_by_user_id",
            "purpose",
            "funding_type_name",
            "activity_type_name",
            "product_group_id",
            "vendor_code",
            "company_code",
            "company_name",
            "t4w_asp",
            "promotion_pricing_amount",
            "discount_per_unit",
            "funding_per_asin",
            "incremental_per_unit",
            "shipped_units",
            "incremental_gains",
        ]

    @staticmethod
    def aggregate_to_asin_level(df):
        # Calculate funding_per_asin before aggregation
        df["funding_per_asin"] = df["coop_amount"] / df["shipped_units"]
        df["funding_per_asin"] = round(df["funding_per_asin"], 2)
        df.drop(columns=["coop_amount"], inplace=True)

        # Group by ASIN level columns and aggregate
        group_cols = DataAggregator.get_group_columns()
        asin_level_df = (
            df.groupby(group_cols).agg({"shipped_units": "sum"}).reset_index()
        )

        return asin_level_df

    @staticmethod
    def create_vendor_agreement_view(df):
        vendor_cols = [
            "region_id",
            "marketplace_key",
            "period",
            "product_group_id",
            "vendor_code",
            "company_code",
            "company_name",
            "funding_per_asin",
            "incremental_gains",
        ]

        vendor_agreement_df = (
            df[vendor_cols]
            .groupby(vendor_cols[:-1])
            .sum("incremental_gains")
            .reset_index()
        )

        vendor_agreement_df.sort_values(
            ["incremental_gains"], ascending=False, inplace=True
        )

        return vendor_agreement_df


class Calculator:
    @staticmethod
    def calculate_incremental_metrics(df):
        # Calculate discount per unit
        df["discount_per_unit"] = df["t4w_asp"] - df["promotion_pricing_amount"]
        df["discount_per_unit"] = np.where(
            df["t4w_asp"].isna(), np.nan, df["discount_per_unit"]
        )
        df["discount_per_unit"] = round(df["discount_per_unit"], 2)

        # Calculate incremental per unit
        df["incremental_per_unit"] = df["funding_per_asin"] - df["discount_per_unit"]
        df["incremental_per_unit"] = np.where(
            (df["incremental_per_unit"] < 0) | (df["incremental_per_unit"].isna()),
            0,
            df["incremental_per_unit"],
        )
        df["incremental_per_unit"] = round(df["incremental_per_unit"], 2)

        # Calculate total incremental gains
        df["incremental_gains"] = df["incremental_per_unit"] * df["shipped_units"]
        df["incremental_gains"] = np.where(
            (df["incremental_gains"] < 0) | (df["incremental_gains"].isna()),
            0,
            df["incremental_gains"],
        )
        df["incremental_gains"] = round(df["incremental_gains"], 2)

        return df

    @staticmethod
    def sort_and_order_columns(df):
        # Sort values
        df = df.sort_values(
            by=["marketplace_key", "region_id", "incremental_gains"],
            ascending=[True, True, False],
        )

        # Reorder columns
        col_order = DataAggregator.get_final_column_order()
        df = df[col_order]

        return df


class ExcelWriter:
    @staticmethod
    def write_to_excel(df, vendor_agreement_df, output_path):
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

        outputs = {"ASIN Level": df, "Vendor-Agreement Level": vendor_agreement_df}

        for sheet_name, data in outputs.items():
            ws = wb.create_sheet(sheet_name)
            rows = dataframe_to_rows(data, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        print("\n" + "*" * 8 + f"  Output saved to {output_path}.  " + "*" * 8 + "\n")


def main():
    # Initialize processor and load data
    processor = DataProcessor()
    df = processor.load_data()

    # Clean data
    cleaner = DataCleaner()
    df = cleaner.convert_dtypes(df)
    df = cleaner.standardize_columns(df)

    # Aggregate to ASIN level first
    aggregator = DataAggregator()
    asin_level_df = aggregator.aggregate_to_asin_level(df)

    # Calculate incremental metrics
    calculator = Calculator()
    asin_level_df = calculator.calculate_incremental_metrics(asin_level_df)
    asin_level_df = calculator.sort_and_order_columns(asin_level_df)

    # Create vendor agreement view
    vendor_agreement_df = aggregator.create_vendor_agreement_view(asin_level_df)

    # Write to Excel
    writer = ExcelWriter()
    writer.write_to_excel(asin_level_df, vendor_agreement_df, processor.output_path)

    print("\n" + "*" * 15 + "  Analysis Complete  " + "*" * 15 + "\n")


if __name__ == "__main__":
    main()
